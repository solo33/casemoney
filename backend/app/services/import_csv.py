"""Import transactions from CSV/XLSX/XLS files.

Expected columns:
    date;account;category;amount;currency;description;transfer

Rules:
- date: dd.mm.yyyy, yyyy-mm-dd, dd/mm/yyyy, or Excel date cell
- amount: negative = expense, positive = income
- category: may contain parent\\child hierarchy
- transfer: optional destination/counterparty account for account transfers
"""
import csv
import io
from html.parser import HTMLParser
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

from app.models.account import Account
from app.models.account_balance import AccountBalance
from app.models.category import Category
from app.models.transaction import Transaction, TransactionType
from app.models.user_currency import UserCurrency
from app.models.user import User


@dataclass
class ParsedRow:
    line_no: int
    date: Optional[str]                   # ISO YYYY-MM-DD
    account: str
    category_path: Optional[str]          # как в CSV: "Покупки\\Подарки" или ""
    category_parent: Optional[str]
    category_child: Optional[str]
    amount: float                         # подписанная сумма (исходная)
    abs_amount: float
    currency: str
    description: Optional[str]
    transfer_to: Optional[str]            # имя счёта-противоположной стороны
    tx_type: str                          # "income" | "expense" | "transfer"
    error: Optional[str] = None


@dataclass
class ImportPreview:
    rows: list[ParsedRow] = field(default_factory=list)
    new_accounts: list[str] = field(default_factory=list)
    existing_accounts: list[str] = field(default_factory=list)
    new_categories: list[dict] = field(default_factory=list)   # [{path, type}]
    existing_categories: list[str] = field(default_factory=list)
    currencies_to_add: list[str] = field(default_factory=list)  # валюты не в user_currencies
    totals: dict = field(default_factory=dict)


COLUMN_ALIASES = {
    "date": "date",
    "дата": "date",
    "account": "account",
    "счет": "account",
    "счёт": "account",
    "category": "category",
    "категория": "category",
    "amount": "amount",
    "total": "amount",
    "сумма": "amount",
    "expense": "expense",
    "расход": "expense",
    "income": "income",
    "доход": "income",
    "currency": "currency",
    "валюта": "currency",
    "record": "record",
    "запись": "record",
    "description": "description",
    "note": "description",
    "comment": "description",
    "описание": "description",
    "комментарий": "description",
    "transfer": "transfer",
    "перевод": "transfer",
}
REQUIRED_COLUMNS = {"date", "account", "amount"}
SPLIT_AMOUNT_COLUMNS = {"date", "account", "expense", "income"}
POSITIONAL_COLUMNS = ["date", "account", "category", "amount", "currency", "description", "transfer"]


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _parse_amount(value) -> float:
    """Преобразует '-600,00' → -600.0"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date(value) -> Optional[str]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = value.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _normalize_header(raw_header: list) -> dict[str, int]:
    normalized = {}
    for idx, value in enumerate(raw_header):
        key = _stringify(value).lower().strip()
        key = COLUMN_ALIASES.get(key, key)
        if key and key not in normalized:
            normalized[key] = idx
    return normalized


def _parse_rows(rows: list[list]) -> list[ParsedRow]:
    if not rows:
        return []

    header_map = _normalize_header(rows[0])
    header_keys = set(header_map)
    has_header = REQUIRED_COLUMNS.issubset(header_keys) or SPLIT_AMOUNT_COLUMNS.issubset(header_keys)
    if not has_header:
        header_map = {name: idx for idx, name in enumerate(POSITIONAL_COLUMNS)}

    parsed: list[ParsedRow] = []
    start = 1 if has_header else 0
    for raw_idx, raw in enumerate(rows[start:], start=start + 1):
        if not raw or all(not _stringify(c) for c in raw):
            continue

        def get(col: str):
            idx = header_map.get(col)
            if idx is None or idx >= len(raw):
                return ""
            return raw[idx]

        date_iso = _parse_date(get("date"))
        account = _stringify(get("account"))
        raw_record = _stringify(get("record"))
        category_path = _stringify(get("category"))
        amount = _parse_amount(get("amount"))
        if not amount and ("expense" in header_map or "income" in header_map):
            expense = _parse_amount(get("expense"))
            income = _parse_amount(get("income"))
            amount = income if income else expense
        currency = (_stringify(get("currency")) or "RUB").upper()
        description = _stringify(get("description")) or None
        transfer_to = _stringify(get("transfer")) or None

        record_lower = raw_record.lower()
        if not category_path and raw_record:
            if record_lower.startswith("расход:"):
                category_path = raw_record.split(":", 1)[1].strip()
            elif record_lower.startswith("доход:"):
                category_path = raw_record.split(":", 1)[1].strip()
        if not transfer_to and record_lower.startswith("перевод"):
            if ":" in raw_record:
                transfer_to = raw_record.split(":", 1)[1].strip()
            else:
                transfer_to = raw_record.replace("Перевод", "").replace("перевод", "").strip()

        if transfer_to:
            tx_type = "transfer"
        elif record_lower.startswith("доход:"):
            tx_type = "income"
            amount = abs(amount)
        elif record_lower.startswith("расход:"):
            tx_type = "expense"
            amount = -abs(amount)
        else:
            tx_type = "expense" if amount < 0 else ("income" if amount > 0 else "expense")

        parent = child = None
        if category_path:
            parts = [p.strip() for p in category_path.split("\\") if p.strip()]
            if len(parts) >= 2:
                parent, child = parts[0], parts[-1]
            elif len(parts) == 1:
                parent = parts[0]

        err = None
        if not date_iso:
            err = "не удалось распарсить дату"
        elif not account:
            err = "пустой счет"

        parsed.append(ParsedRow(
            line_no=raw_idx,
            date=date_iso,
            account=account,
            category_path=category_path or None,
            category_parent=parent,
            category_child=child,
            amount=amount,
            abs_amount=abs(amount),
            currency=currency,
            description=description,
            transfer_to=transfer_to,
            tx_type=tx_type,
            error=err,
        ))
    return parsed


def parse_csv(content: bytes) -> list[ParsedRow]:
    """Parse raw CSV content into ParsedRow objects."""
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Не удалось распознать кодировку файла")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    return _parse_rows(rows)


def parse_xlsx(content: bytes) -> list[ParsedRow]:
    """Parse XLSX content using the first worksheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError("Для импорта XLSX установите зависимость openpyxl") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    return _parse_rows(rows)


def parse_xls(content: bytes) -> list[ParsedRow]:
    """Parse legacy XLS content using the first worksheet."""
    prefix = content[:256].lstrip().lower()
    if prefix.startswith(b"<") or prefix.startswith(b"\xef\xbb\xbf<"):
        return parse_html_table(content)

    try:
        import xlrd
    except ImportError as e:
        raise ValueError("Для импорта XLS установите зависимость xlrd") from e

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = []
    for row_idx in range(sheet.nrows):
        row = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = datetime(*xlrd.xldate_as_tuple(value, book.datemode))
            row.append(value)
        rows.append(row)
    return _parse_rows(rows)


def parse_html_table(content: bytes) -> list[ParsedRow]:
    """Parse bank exports saved as .xls but containing an HTML table."""
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Не удалось распознать кодировку HTML-таблицы")

    parser = _SimpleTableParser()
    parser.feed(text)
    rows = parser.rows
    if not rows:
        return []
    return _parse_rows(rows)


class _SimpleTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
            self._in_cell = True

    def handle_data(self, data):
        if self._in_cell and self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._row is not None and self._cell is not None:
            value = " ".join("".join(self._cell).split())
            self._row.append(value)
            self._cell = None
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            if any(str(c).strip() for c in self._row):
                self.rows.append(self._row)
            self._row = None


def parse_file(filename: str, content: bytes) -> list[ParsedRow]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv" or not suffix:
        return parse_csv(content)
    if suffix == ".xlsx":
        return parse_xlsx(content)
    if suffix == ".xls":
        return parse_xls(content)
    raise ValueError("Поддерживаются файлы CSV, XLSX и XLS")


def build_preview(db: Session, user_id: int, rows: list[ParsedRow]) -> ImportPreview:
    """Собирает сводку: что нового нужно создать, итоги."""
    existing_accounts = {
        a.name: a for a in db.query(Account).filter(Account.user_id == user_id).all()
    }
    all_existing_cats = db.query(Category).filter(Category.user_id == user_id).all()
    cats_by_id_pre = {c.id: c for c in all_existing_cats}
    existing_categories: dict[tuple, Category] = {}
    for c in all_existing_cats:
        parent_name = None
        if c.parent_id and c.parent_id in cats_by_id_pre:
            parent_name = cats_by_id_pre[c.parent_id].name.lower()
        existing_categories[(parent_name, c.name.lower())] = c
    existing_user_currencies = {
        uc.currency.upper() for uc in db.query(UserCurrency).filter(UserCurrency.user_id == user_id).all()
    }

    new_accounts: set[str] = set()
    seen_accounts: set[str] = set()
    new_categories: dict[tuple, str] = {}   # (parent_name, name) → type
    seen_categories: set[str] = set()
    currencies: set[str] = set()
    total_income = total_expense = 0.0
    transfer_count = ok_count = err_count = 0

    for r in rows:
        if r.error:
            err_count += 1
            continue
        ok_count += 1

        currencies.add(r.currency)
        if r.account:
            if r.account in existing_accounts:
                seen_accounts.add(r.account)
            else:
                new_accounts.add(r.account)
        # Transfer counterpart account
        if r.transfer_to:
            if r.transfer_to in existing_accounts:
                seen_accounts.add(r.transfer_to)
            else:
                new_accounts.add(r.transfer_to)
            transfer_count += 1
        else:
            if r.amount < 0:
                total_expense += r.abs_amount
            else:
                total_income += r.abs_amount

        # Категории (только для не-transfer)
        if not r.transfer_to:
            cat_type = "expense" if r.amount < 0 else "income"
            # Корневая
            if r.category_parent:
                key = (None, r.category_parent.lower())
                if key in existing_categories:
                    seen_categories.add(r.category_parent)
                else:
                    new_categories.setdefault(key, cat_type)
            # Дочерняя
            if r.category_child:
                key = (r.category_parent.lower() if r.category_parent else None,
                       r.category_child.lower())
                if key in existing_categories:
                    seen_categories.add(r.category_child)
                else:
                    new_categories.setdefault(key, cat_type)

    p = ImportPreview()
    p.rows = rows
    p.new_accounts = sorted(new_accounts)
    p.existing_accounts = sorted(seen_accounts)
    p.new_categories = [
        {"name": (k[1] if not k[0] else f"{k[0]}\\{k[1]}"), "type": t}
        for k, t in sorted(new_categories.items(), key=lambda item: ((item[0][0] or ""), item[0][1]))
    ]
    p.existing_categories = sorted(seen_categories)
    p.currencies_to_add = sorted(currencies - existing_user_currencies)
    p.totals = {
        "rows_total": len(rows),
        "ok": ok_count,
        "errors": err_count,
        "transfers": transfer_count,
        "income_sum": round(total_income, 2),
        "expense_sum": round(total_expense, 2),
    }
    return p


def execute_import(db: Session, user_id: int, rows: list[ParsedRow]) -> dict:
    """Создаёт сущности и транзакции в БД. Возвращает счётчики."""
    main_currency = "RUB"
    user = db.query(User).filter(User.id == user_id).first()
    if user and user.main_currency:
        main_currency = user.main_currency.upper()

    # Кэш существующих
    accounts_cache: dict[str, Account] = {
        a.name: a for a in db.query(Account).filter(Account.user_id == user_id).all()
    }
    # Кэш категорий по ключу (parent_name_lower_or_None, name_lower) — чтобы поддержать
    # одноимённые подкатегории под разными родителями (напр. Развлечения в Отдых и в Личные Фима).
    all_cats = db.query(Category).filter(Category.user_id == user_id).all()
    cats_by_id = {c.id: c for c in all_cats}
    categories_cache: dict[tuple, Category] = {}
    for c in all_cats:
        parent_name = None
        if c.parent_id and c.parent_id in cats_by_id:
            parent_name = cats_by_id[c.parent_id].name.lower()
        categories_cache[(parent_name, c.name.lower())] = c
    user_currencies_cache: set[str] = {
        uc.currency.upper() for uc in db.query(UserCurrency).filter(UserCurrency.user_id == user_id).all()
    }

    def ensure_user_currency(currency: str):
        c = currency.upper()
        if c in user_currencies_cache:
            return
        db.add(UserCurrency(user_id=user_id, currency=c, auto=True))
        user_currencies_cache.add(c)

    def ensure_account(name: str) -> Account:
        if name in accounts_cache:
            return accounts_cache[name]
        acc = Account(name=name, type="cash", user_id=user_id)
        db.add(acc)
        db.flush()
        accounts_cache[name] = acc
        return acc

    def ensure_balance(account: Account, currency: str) -> AccountBalance:
        currency = currency.upper()
        for b in account.balances:
            if b.currency == currency:
                return b
        bal = AccountBalance(account_id=account.id, currency=currency, balance=0.0)
        db.add(bal)
        db.flush()
        account.balances.append(bal)
        return bal

    def ensure_category(name: str, cat_type: str, parent: Optional[Category] = None) -> Category:
        parent_name = parent.name.lower() if parent else None
        key = (parent_name, name.lower())
        if key in categories_cache:
            return categories_cache[key]
        # Если у корня (parent=None) уже есть категория с таким именем но другим parent — НЕ
        # переиспользуем, создаём новую. Так "Развлечения" в Отдых и в Личные Фима — это разные.
        cat = Category(
            user_id=user_id,
            name=name,
            type=cat_type,
            color="#9f1239",
            icon=None,
            parent_id=parent.id if parent else None,
        )
        db.add(cat)
        db.flush()
        categories_cache[key] = cat
        return cat

    imported = 0
    skipped = 0
    errors: list[dict] = []

    for r in rows:
        if r.error:
            skipped += 1
            continue
        try:
            ensure_user_currency(r.currency)
            account = ensure_account(r.account)
            bal = ensure_balance(account, r.currency)

            # Категория
            cat: Optional[Category] = None
            if not r.transfer_to:
                cat_type = "expense" if r.amount < 0 else "income"
                parent_cat = None
                if r.category_parent:
                    parent_cat = ensure_category(r.category_parent, cat_type)
                if r.category_child:
                    cat = ensure_category(r.category_child, cat_type, parent=parent_cat)
                elif parent_cat:
                    cat = parent_cat

            # Тип транзакции
            if r.transfer_to:
                tx_type = TransactionType.transfer
            else:
                tx_type = TransactionType.expense if r.amount < 0 else TransactionType.income

            description = r.description
            if r.transfer_to:
                arrow = "->" if r.amount < 0 else "<-"
                prefix = f"Перевод {arrow} {r.transfer_to}"
                description = f"{prefix}: {description}" if description else prefix

            tx_date = datetime.strptime(r.date, "%Y-%m-%d") if r.date else None

            to_account_id = to_amount = to_currency = None
            if r.transfer_to:
                other = ensure_account(r.transfer_to)
                ensure_balance(other, r.currency)
                if r.amount < 0:
                    source = account
                    target = other
                else:
                    source = other
                    target = account
                account = source
                bal = ensure_balance(source, r.currency)
                to_account_id = target.id
                to_amount = r.abs_amount
                to_currency = r.currency.upper()

            tx = Transaction(
                amount=r.abs_amount,
                currency=r.currency.upper(),
                type=tx_type,
                description=description,
                date=tx_date,
                account_id=account.id,
                category_id=cat.id if cat else None,
                user_id=user_id,
                to_account_id=to_account_id,
                to_amount=to_amount,
                to_currency=to_currency,
            )
            db.add(tx)

            if tx_type == TransactionType.transfer:
                bal.balance -= r.abs_amount
                target_bal = ensure_balance(target, r.currency)
                target_bal.balance += r.abs_amount
            elif r.amount < 0:
                bal.balance -= r.abs_amount
            else:
                bal.balance += r.abs_amount

            imported += 1
        except Exception as e:
            skipped += 1
            errors.append({"line_no": r.line_no, "error": str(e)})

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}
